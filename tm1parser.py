import openpyxl
import os
import re
import copy
import xlwings as xw


# todo add description
class TM1PaxParser:

    def __init__(self, filepath, application_name='', dfname_pattern='TM1RPTVIEWRNG', cell_pattern='VIEW'):
        self.filepath = filepath
        self.application_name = application_name
        self.workbook = openpyxl.load_workbook(self.filepath)
        self.defined_names = self.workbook.defined_names.definedName
        self.dfname_pattern = dfname_pattern
        self.cell_pattern = cell_pattern
        self.cube_names = []
        self.sheet_names = self.workbook.sheetnames
        self.structure = []
        self.xlsm_file = './get_TM1Buttons.xlsm'
        self.buttons = {}

    def add_CubeName(self, cube_name, sheet_name):
        try:
            process = self.buttons[sheet_name]
        except KeyError:
            process = []

        app_structure = {
            'app_name': self.application_name,
            'app_path': self.filepath,
            'view_name': sheet_name,
            'cube_name': cube_name,
            'proc_name': process
        }
        self.structure.append(app_structure)
        self.cube_names.append(cube_name)

    def parse(self):
        self.parse_DefinedNames()
        if len(self.cube_names) == 0:
            self.parse_CellRange()

    def parse_DefinedNames(self):
        """
        find cube name in Excel Defined Names by pattern.
        """
        self.find_Buttons()
        for defined_name in self.defined_names:
            if self.dfname_pattern in defined_name.name:
                sheet_name, cell_range = defined_name.attr_text.split('!')
                sheet_name = sheet_name.replace('\'', '')
                cell_ = cell_range.split(':')[0].replace('$', '')
                cube_name = self.workbook[sheet_name][cell_].value

                self.add_CubeName(cube_name, sheet_name)

    def parse_CellRange(self):
        for sheet in self.sheet_names:
            ws = self.workbook[sheet]
            for row in range(1, 6):
                for col in range(1, 10):
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, str):
                        if self.cell_pattern in cell_value:
                            cube_name = cell_value.split(',')[0].split(':')[1].replace('"', '')
                            self.add_CubeName(cube_name, sheet)

    def find_Buttons(self):
        from multiprocessing.connection import Client
        try:
            xlx_server = Client(('localhost', 6000))
        except ConnectionRefusedError:
            #TODO add logging
            pass
        else:
            xlx_server.send(self.filepath)
            self.buttons = xlx_server.recv()
            xlx_server.close()

        # if os.path.exists(self.xlsm_file):
        #     xlx_app = xw.App(visible=False)
        #     wb = xlx_app.books.open(self.xlsm_file)
        #     wb.sheets['Sheet1'].range('A1').value = self.filepath
        #     vba = wb.app.macro('get_Buttons')
        #     vba()
        #
        #     fill = 1
        #     CurRange = wb.sheets['Sheet1'].range('A2:B2')
        #     while fill:
        #         text = CurRange.value
        #         if text[0] is None:
        #             fill = 0
        #         else:
        #             key_, val_ = text
        #             if val_ is None:
        #                 val_ = ''
        #             buttons = val_.split(',')[:-1]
        #             self.buttons[key_] = buttons
        #             CurRange = CurRange.offset(1)
        #     wb.close()
        #     xlx_app.quit()


class TM1AppParser:

    def __init__(self, app_dir, df_dir, db_pattern, dimensions, description, encoding='utf-8'):
        self.app_dir = app_dir
        self.structure = {}
        self.datafiles_dir = df_dir
        self.db_pattern = db_pattern
        self.dimensions = self.import_Dimensions(dimensions)
        self.description = description
        self.desc_pro = self.import_Description(self.description['process'])
        self.desc_cube = self.import_Description(self.description['cube'])
        self.desc_dim = self.import_Description(self.description['dimension'])
        self.start_position = self.calc_StartPosotion()
        # self.cur_structure = self.structure
        self.cur_folder = ''
        self.cur_AppName = ''
        self.cur_ViewName = ''
        self.cur_CubeName = ''
        self.cur_AppPath = ''
        self.cur_AppType = ''
        self.cur_FilePath = ''
        self.encode = encoding

    def import_Dimensions(self, path, encoding='cp1251'):
        dict_dim = {}

        if os.path.exists(path):
            with open(path, 'r', encoding=encoding) as f:
                for line in f:
                    #todo fix variable count after fix tm1process
                    key, value, dim_num = re.split(',', line.replace('"', ''), maxsplit=2)
                    try:
                        dict_dim[key].append(value)
                    except KeyError:
                        dict_dim[key] = []
                        dict_dim[key].append(value)

        return dict_dim

    def import_Description(self, path, encoding='cp1251'):
        dict_desc = {}

        if os.path.exists(path):
            with open(path, 'r', encoding=encoding) as f:
                for line in f:
                    key, value = re.split(',', line.replace('"', ''), maxsplit=1)
                    dict_desc[key] = value

        return dict_desc

    def calc_StartPosotion(self):
        return len(os.path.split(self.app_dir)[0]) + 1

    def clean_CurVars(self):
        self.cur_AppName = ''
        self.cur_ViewName = ''
        self.cur_CubeName = ''
        self.cur_AppPath = ''
        self.cur_AppType = ''
        self.cur_FilePath = ''

    def fit(self):
        folders = sorted(list(os.walk(self.app_dir, topdown=True)))
        for folder in folders:
            folderpath, _, filenames = folder
            # TODO delete or not?
            # filenames.sort()

            self.cur_structure = self.structure
            if '}STC' in folderpath:
                continue
            *parent_folders, self.cur_folder = folderpath[self.start_position:].split('/')

            for parent_folder in parent_folders:
                self.cur_structure = self.cur_structure[parent_folder]['subfolder']
            self.cur_structure[self.cur_folder] = {'subfolder': {}, 'app': []}

            for filename in sorted(filenames):
                self.clean_CurVars()
                self.cur_AppName, self.cur_AppType = os.path.splitext(filename)
                if len(self.cur_AppType) > 0:
                    # todo create new finction get_filepath and get_info
                    self.cur_AppPath = self.get_filepath(self, folderpath, filename)
                    with open(self.cur_AppPath, 'r', encoding='utf-8') as f:
                        for line in f:
                            search_string = re.search('ENTRYREFERENCE=TM1:', line)
                            if search_string is not None:
                                line_parts = line.replace('\n', '').split('/')
                                if self.cur_AppType == '.view':
                                    self.cur_FilePath = line_parts[-2:]
                                elif self.cur_AppType == '.blob':
                                    self.cur_FilePath = line_parts[-1].replace('\\', '/')

                    if self.cur_AppType == '.view':
                        self.cur_CubeName, self.cur_ViewName = self.cur_FilePath
                        self.create_AppStructure()

                    elif self.cur_AppType == '.blob':
                        xlx_path = os.path.join(self.datafiles_dir, self.cur_FilePath)
                        parser = TM1PaxParser(xlx_path)
                        parser.parse()
                        # TODO rename 'app'
                        for app in parser.structure:
                            self.cur_CubeName = app['cube_name']
                            self.create_AppStructure()

                    else:
                        self.create_AppStructure()

    @staticmethod
    def get_filepath(self, dirpath, filename):
        filepath = os.path.join(dirpath, filename)
        # в Windows модет возникнуть ошибка из-за слишком длинного имени
        # только для этой ОС исправляем путь - добавляем \\?\
        if os.name == 'nt':
            if filepath[:2] == '\\\\':
                # UNC - используется для сетевых папок
                rplc = u'\\\\?\\UNC\\'
            else:
                rplc = u'\\\\?\\' + filepath[:2]
            filepath = rplc + filepath[2:]
        return filepath


    def create_AppStructure(self):
        cube_path = os.path.join(self.datafiles_dir, self.cur_CubeName)
        parser = TM1CubeParser(cube_path)
        parser.parse_rux(self.datafiles_dir, self.db_pattern)

        # todo add 'description' to structure
        file_structure = parser.structure
        file_structure['app_name'] = self.cur_AppName
        file_structure['app_path'] = self.cur_AppPath
        # file_structure['proc_name'] = parser.process
        file_structure['view_name'] = self.cur_ViewName
        file_structure['dim_names'] = self.add_Dimensions(file_structure)
        self.add_Description(file_structure)

        self.cur_structure[self.cur_folder]['app'].append(
            {'name': self.cur_AppName, 'type': self.cur_AppType, 'info': file_structure}
        )

    def add_Dimensions(self, file_structure):
        try:
            return self.dimensions[file_structure['cube_name']]
        except KeyError:
            return []

    def find_Descripion(self, value, description):
        try:
            new_value = (value, description[value])
        except KeyError:
            new_value = (value, '')
        return new_value

    def add_Description(self, sructure):
        for name, desc in zip(('cube_name', 'proc_name', 'dim_names'), (self.desc_cube, self.desc_pro, self.desc_dim)):
            structure_value = sructure[name]
            if isinstance(structure_value, str):
                sructure[name] = self.find_Descripion(structure_value, desc)
            elif isinstance(structure_value, list):
                sructure[name] = [self.find_Descripion(value, desc) for value in structure_value]


class TM1CubeParser:

    def __init__(self, filepath, encoding='utf-8'):
        self.filepath = filepath
        self.name = os.path.basename(self.filepath).split('.')[0]
        self.sources = set()
        self.destinations = set()
        self.rux_exists = False
        self.drill_exists = False
        self.process = set()
        self.dimensions = []
        self.structure = {'cube_name': self.name, 'rux': {}, 'dim_names': [], 'proc_name': set()}
        self.file_encode = encoding

    def get_CodeLines(self, fullpath):
        if os.path.exists(fullpath):
            with open(fullpath, 'r', encoding=self.file_encode) as f:
                lines = f.read()
        else:
            lines = []

        return lines

    def update_structure(self):
        self.structure = {
            'cube_name': self.name,
            'dim_names': self.dimensions,
            'proc_name': self.process,
            'rux': {
                'src_name': self.sources,
                'dst_name': self.destinations
            }
        }


    def parse_rux(self, datafiles, db_ptrn, suffix='.rux'):
        """
        To find in RUX files destination and source cubes
        """
        rux_path = os.path.join(datafiles, f'{self.name}{suffix}')
        codelines = self.get_CodeLines(rux_path)
        if len(codelines) > 0:
            self.find_SourceName(codelines, db_ptrn, storages=(self.sources, self.destinations))
            self.structure['rux']['src_name'] = self.sources
            self.structure['rux']['dst_name'] = self.destinations

        drillpath = os.path.join(datafiles, f'}}CubeDrill_{self.name}{suffix}')
        codelines = self.get_CodeLines(drillpath)
        if len(codelines) > 0:
            self.find_SourceName(codelines, ptrn=r'[\"\'].*[\"\']', storages=(self.process, ), split_operators=False)
            self.structure['proc_name'] = self.process

    # todo rename function
    def find_SourceName(self, codelines, ptrn, storages, split_operators=True):
        codelines = self.CodePreparation(codelines)
        raw_srccubes, raw_dstcubes = self.CodeSplit(codelines)

        # todo transfer this part to another function
        for raw_srccube in raw_srccubes:
            if isinstance(raw_srccube, str):
                if split_operators:
                    cleand_srccubes = self.split_operator(raw_srccube)
                else:
                    cleand_srccubes = raw_srccube.split(',')

                for cleand_srccube in cleand_srccubes:
                    res = re.search(ptrn, cleand_srccube)
                    if res is not None:
                        storages[0].add(self.del_qoutes(res.group(0)))

        for raw_dstcube in raw_dstcubes:
            res = re.search(ptrn, raw_dstcube)
            if res is not None:
                storages[1].add(res.group(0))

    def CodePreparation(self, codelines):
        prep_codelines = re.sub(r'\s*#.*\n', '', codelines).replace('\n', '').replace('\t', '').split(';')
        prep_codelines = list(map(str.strip, prep_codelines))

        return prep_codelines

    def CodeSplit(self, codelines):
        IF_ptrn = r'(?<=IF)\s*\(.*'
        OTHER_ptrn = r'=\s*[SNC]?\s*:(.*)'
        FEED_ptrn = r'(?<=\=)\s*>.*'

        src, dst = [], []
        if_formula = []
        other_formula = []
        feed_flg = 0

        for codeline in codelines:

            fmt_line = codeline.upper().strip()
            if fmt_line == '':  # пропускаем пустые сроки
                pass
            elif fmt_line[0] == '#':  # пропускаем заакомментированные строки
                pass
            elif 'FEEDERS' in fmt_line:  # после этой строки кубов-источников нет
                feed_flg = 1
            else:
                if feed_flg == 0:
                    if_match = re.search(IF_ptrn, codeline)  # проверка конструкции IF
                    other_match = re.search(OTHER_ptrn, codeline)

                    if if_match is not None:
                        formula = if_match.group(0).strip()[1:]
                        if_formula.append(formula)
                    elif other_match is not None:
                        formula = other_match.group(1)
                        other_formula.append(formula)
                else:
                    feed_match = re.search(FEED_ptrn, codeline)
                    if feed_match is not None:
                        formula = feed_match.group(0).strip().replace('>', '')
                        dst_cubes = re.split('(DB\(.*)', formula)
                        dst += dst_cubes

        if len(if_formula) == 0:
            src = other_formula
        else:
            if_src = []
            for formula in if_formula:
                _, *parts = self.split_if(formula)
                for part in parts:
                    src_part, dst_part = self.CodeSplit(part)
                    if len(src_part) > 0:
                        if_src.append(src_part)
                    elif len(dst_part) > 0:
                        #TODO need replace to dst_part?
                        if_src.append(src_part)
                    else:
                        if_src.append(part)

            src.extend(if_src)
            src.extend(other_formula)

        return src, dst

    def split_if(self, formula):
        delimiter, start_ = 0, 0
        if_part, if_del = 0, []

        for num, sign in enumerate(formula):
            if sign == '(':
                delimiter += 1
            if sign == ')':
                delimiter -= 1

            if delimiter == 0 and sign == ',':
                if_del.append(formula[start_:num])
                start_ = num + 1
                if_part += 1
                if if_part == 2:
                    if_del.append(formula[start_:])
                    break
        return if_del

    def split_operator(self, codeline):
        operators = ['+', '-', '*', '/']
        rep_sign = '$_$'
        fmt_line = codeline

        if isinstance(codeline, str):
            for operator in operators:
                fmt_line = fmt_line.replace(operator, rep_sign)

            arr = fmt_line.split(rep_sign)

            return arr

    def del_qoutes(self, string):
        string = string.strip()
        for chr in ['\"', "'", ';']:
            string = string.replace(chr, '')
        return string


class TM1ProcessParser:

    def __init__(self, filepath, tech_dict, datasources, functions, structure={}, encoding='utf-8'):
        self.filepath = filepath
        self.tech_dict = tech_dict
        self.datasources = datasources
        self.functions = functions
        self.file_encode = encoding
        self.line_num = 0
        self.change_src = 0
        self.code_board_flg = 0
        self.code_start_line = 0
        self.code_end_line = 0
        self.code_part = ''
        self.src_function = ''
        self.join_line = ''
        self.variables = {}
        # todo in previous code it imported from globvar
        self.structure = copy.deepcopy(structure)
        self.file_exists = os.path.exists(self.filepath)
        self.chr_replace = ('\n', '\t')
        self.errors_msg = []

    def get_CodeLines(self):
        if self.file_exists:
            with open(self.filepath, 'r', encoding=self.file_encode) as f:
                lines = f.readlines()
        else:
            lines = []

        return lines

    def clean_Line(self, line):
        for char in self.chr_replace:
            line = line.replace(char, '')

        return line

    def parse(self):
        lines = self.get_CodeLines()
        self.structure['pro_name'] = os.path.basename(self.filepath).split('.')[0]
        for line in lines:
            self.line_num += 1
            fmt_line = self.clean_Line(line)
            if fmt_line == '':
                continue
            elif fmt_line[0] == '#':
                continue

            if self.line_num == self.code_end_line:
                self.code_board_flg = 0
                self.code_part = ''

            # todo create function
            if self.code_part in self.tech_dict['code'][:4]:
                if ';' not in fmt_line:
                    self.join_line += fmt_line.strip()
                    continue
                else:
                    self.join_line = ''
                    fmt_line = self.join_line + fmt_line.strip().replace(';', '')
            else:
                # todo replace line to fmt_line?
                fmt_line = line.replace(';', '')

            if self.code_board_flg == 0:
                tech_code, *tech_ = fmt_line.split(',')
                tech_value = str.join('', tech_)

                if tech_code.isnumeric():
                    if tech_code in self.tech_dict.keys():
                        dict_key = self.tech_dict[tech_code]
                        self.structure[dict_key] = self.del_qoutes(tech_value)
                    elif tech_code in self.tech_dict['code']:
                        self.code_start_line = self.line_num + 1
                        self.code_end_line = int(tech_value) + self.code_start_line
                        self.code_board_flg = 1
                        self.code_part = tech_code
                continue
            else:
                fmt_lineLow = fmt_line.lower()
                if self.code_part == '590':
                    var, val = re.split(',', fmt_line, maxsplit=1)
                    self.variables[var.lower()] = self.del_qoutes(val)
                    continue
                elif self.code_part == '572':
                    if re.search(r'datasourcetype\s*=', fmt_lineLow) is not None:
                        val = self.del_qoutes(line.split('=')[1])
                        self.structure['type_source'] = val
                        self.src_function = self.datasources[val.upper().strip()]
                        self.structure['source'] = ''
                        self.change_src = 1
                    elif self.src_function in fmt_lineLow and self.change_src == 1:
                        src = fmt_line.split('=')[1].strip()

                        if src[0] == src[-1] and src[0] in ['"', '\'']:
                            val = src[1:-1]
                        else:
                            val = self.variables[src.lower()]
                        self.structure['source'] = self.del_qoutes(val)
                        self.change_src = 0

            try:
                if '=' in fmt_line:
                    if 'if' in fmt_lineLow:
                        continue
                    var, val = re.split('=', fmt_line, maxsplit=1)
                    var = var.lower().strip()
                    if '|' in val:
                        val = self.join_string(val)
                    self.variables[var] = self.del_qoutes(val)
            except ValueError:
                error_msg = f'ValueError: {fmt_line}.\n{self.code_board_flg}, {self.code_start_line}, {self.code_end_line}, {self.line_num}'
                self.errors_msg.append(error_msg)
                continue

            for ptrn_func in self.functions.keys():
                if ptrn_func in fmt_lineLow:
                    if re.search(f'textoutput.*{ptrn_func}', fmt_lineLow) is not None:
                        continue
                    dst_type = self.functions[ptrn_func]

                    if dst_type == 'cubes':
                        try:
                            start_point = re.search(f'{ptrn_func}[n,s]\s*\(', fmt_lineLow).span()[1]
                        except AttributeError:
                            continue
                        # fmt_lineSplit = fmt_line[start_point:]
                        fmt_lineSplit = re.split(',', fmt_line[start_point:])
                        if '(' in fmt_lineSplit[0]:
                            r_brackets = re.findall('\(', fmt_lineSplit[0])
                            l_brackets = re.findall('\)', fmt_lineSplit[0])
                            if len(r_brackets) == len(l_brackets):
                                dst = fmt_lineSplit[1].strip()
                            else:
                                exp_flg = 1
                                for ind, splt in enumerate(fmt_lineSplit[1:]):
                                    if exp_flg == 0:
                                        dst = splt.strip()
                                        break
                                    else:
                                        if '(' in splt:
                                            exp_flg += 1
                                        if ')' in splt:
                                            exp_flg -= 1
                        else:
                            dst = fmt_lineSplit[1].strip()
                    elif dst_type == 'dimensions':
                        start_point = re.search(f'{ptrn_func}', fmt_lineLow).span()[1]
                        fmt_line = line[start_point:]
                        dst = fmt_line.split(',')[0].split('(')[1].strip()

                    if dst != '':
                        if dst[0] == dst[-1] and dst[0] in ['"', '\'']:
                            val = self.del_qoutes(dst)
                        else:
                            dst = dst.lower()
                            try:
                                if '|' in dst:
                                    val = self.join_string(dst)
                                else:
                                    val = self.del_qoutes(self.variables[dst])
                            except KeyError:
                                error_msg = f'В процессе "{self.filepath}" переменная "{dst}" не опеределена. Строка - {self.line_num}'
                                self.errors_msg.append(error_msg)
                                continue

                        if val not in self.structure['destination'][dst_type]:
                            self.structure['destination'][dst_type].append(val)
                    break

            if 'executeprocess' in fmt_lineLow:
                d_proc = fmt_line.split(',')[0].split('(')[1].replace(')', '').strip()
                if '|' in d_proc:
                    d_proc = self.join_string(d_proc)
                else:
                    d_proc = self.del_qoutes(d_proc)

                if d_proc not in self.structure['daughter_proc']:
                    self.structure['daughter_proc'].append(d_proc)

            source_type = self.structure['type_source']
            source_name = self.structure['source']
            self.structure['form_source'] = f'{source_type} : {source_name}'

    def join_string(self, string):
        new_string = ''
        for string_part in string.split('|'):
            string_part_low = string_part.lower().strip()
            if string_part_low in self.variables.keys():
                new_string += self.variables[string_part_low]
            else:
                new_string += self.del_qoutes(string_part)
        return new_string

    def del_qoutes(self, string):
        string = string.strip()
        for chr in ['\"', "'", ';']:
            string = string.replace(chr, '')
        return string


if __name__ == '__main__':

    from pprint import pprint
    from glob_var import *

    xlx_path = '/Users/franc/PycharmProjects/tm1find/test/DataFiles/}Externals/08 Выходная форма EXCEL.xlsx_20190917134832.xlsx'
    parser = TM1PaxParser(xlx_path)
    parser.parse()
    print(parser.structure)
    #
    # cub_path = '/Users/franc/PycharmProjects/tm1find/test/DataFiles/ОПР_БДР.RUX'
    # CubeParser = TM1CubeParser(cub_path)
    # CubeParser.parse_rux(DF_DIR, DB_PTRN)
    # # print(CubeParser.sources)
    # pprint(CubeParser.structure)
    # # print(CubeParser.get_structure())

    # description = {
    #     'dimension': DESC_DIM,
    #     'cube': DESC_CUBE,
    #     'process': DESC_PRO
    # }
    #
    # AppParser = TM1AppParser(
    #     app_dir=APP_DIR, df_dir=DF_DIR, db_pattern=DB_PTRN,
    #     dimensions=DIM_FILE, description=description
    # )
    # AppParser.fit()
    # pprint(AppParser.structure)

    # proc_path = '/Users/franc/PycharmProjects/tm1find/test/DataFiles/}Drill_2 Утвержден.pro'
    # parser = TM1ProcessParser(proc_path, tech_dict=tech_dict, datasources=data_sourses, functions=functions, structure=pro_dict_temp)
    # parser.parse()
    # pprint(parser.structure)

